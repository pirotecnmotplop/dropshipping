from __future__ import annotations

import argparse
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

WORKBOOK_NAME = "planejamento_dropshipping_brasil_melhorado.xlsx"
PREVIEW_NAME = "dashboard_preview.png"

REQUIRED_SHEETS = [
    "00 Escolha_Principal",
    "01 Visao_Geral",
    "02 Modelo_Negocio",
    "03 Nichos",
    "04 Produtos",
    "05 Fornecedores",
    "06 Custos_Iniciais",
    "07 Precificacao",
    "08 Loja",
    "09 Anuncios_Pagos",
    "10 Trafego_Organico",
    "11 Cronograma_30_Dias",
    "12 Riscos",
    "13 Legal_Brasil",
    "14 Ferramentas",
    "15 Financeiro_3_Meses",
    "16 Checklist_Final",
]


@dataclass(frozen=True)
class Palette:
    navy: str = "17365D"
    blue: str = "1F4E78"
    sky: str = "D9EAF7"
    ice: str = "EAF2F8"
    green: str = "2E7D32"
    green_light: str = "D9EAD3"
    amber: str = "F6B26B"
    amber_light: str = "FFF2CC"
    red: str = "C00000"
    red_light: str = "F4CCCC"
    purple: str = "674EA7"
    gray: str = "666666"
    gray_light: str = "F3F6F8"
    white: str = "FFFFFF"
    border: str = "D9E2EC"
    text: str = "1F2933"


PALETTE = Palette()

THIN_BORDER = Border(
    left=Side(style="thin", color=PALETTE.border),
    right=Side(style="thin", color=PALETTE.border),
    top=Side(style="thin", color=PALETTE.border),
    bottom=Side(style="thin", color=PALETTE.border),
)
HEADER_FILL = PatternFill("solid", fgColor=PALETTE.blue)
SUBHEADER_FILL = PatternFill("solid", fgColor=PALETTE.sky)
INPUT_FILL = PatternFill("solid", fgColor=PALETTE.ice)
CALC_FILL = PatternFill("solid", fgColor=PALETTE.amber_light)
NOTE_FILL = PatternFill("solid", fgColor=PALETTE.gray_light)
GOOD_FILL = PatternFill("solid", fgColor=PALETTE.green_light)
WARN_FILL = PatternFill("solid", fgColor=PALETTE.amber_light)
BAD_FILL = PatternFill("solid", fgColor=PALETTE.red_light)

CURRENCY_FORMAT = 'R$ #,##0.00'
PERCENT_FORMAT = "0.0%"
INTEGER_FORMAT = "#,##0"
ROI_FORMAT = '0.00"x"'


NICHOS = [
    [
        1,
        "Casa, organização e utilidades práticas",
        9.7,
        "Alto",
        "Média",
        "R$ 40 a R$ 120",
        "Alta",
        "Baixo",
        "Organizador de cabos, suporte adesivo, organizador de geladeira",
        "Pessoas que querem praticidade em casa, escritório ou pequenos espaços",
        "Sim",
        "Produtos simples, visuais, fáceis de demonstrar e com baixo risco técnico.",
    ],
    [
        2,
        "Pet utilitário",
        8.8,
        "Alto",
        "Média",
        "R$ 35 a R$ 100",
        "Alta",
        "Baixo",
        "Escova removedora de pelos, tapete higiênico, bebedouro portátil",
        "Tutores de pets que buscam higiene, passeio e organização",
        "Sim",
        "Boa resposta em vídeos curtos; evitar ração, remédios e promessas veterinárias.",
    ],
    [
        3,
        "Cozinha prática",
        8.5,
        "Médio/alto",
        "Média",
        "R$ 30 a R$ 90",
        "Alta",
        "Médio",
        "Cortador prático, porta-temperos, escorredor dobrável",
        "Famílias, pessoas que cozinham e pequenos apartamentos",
        "Sim",
        "Problemas cotidianos são claros; evitar itens frágeis ou elétricos complexos.",
    ],
    [
        4,
        "Automotivo leve",
        8.3,
        "Médio/alto",
        "Média",
        "R$ 35 a R$ 110",
        "Média/alta",
        "Médio",
        "Suporte veicular, lixeira dobrável, organizador de banco",
        "Motoristas, motoristas de aplicativo e famílias",
        "Sim",
        "Funciona bem com antes/depois; evitar peças mecânicas e instalação técnica.",
    ],
    [
        5,
        "Fitness leve e bem-estar sem promessa médica",
        7.6,
        "Médio",
        "Alta",
        "R$ 30 a R$ 90",
        "Média",
        "Médio",
        "Garrafas, bolsas, elásticos e organizadores de treino",
        "Pessoas que treinam em casa ou buscam rotina mais organizada",
        "Parcial",
        "Pode vender, mas exige cuidado para não prometer saúde, cura ou emagrecimento.",
    ],
    [
        6,
        "Beleza não regulada",
        7.4,
        "Médio",
        "Alta",
        "R$ 25 a R$ 90",
        "Média",
        "Médio",
        "Organizador de maquiagem, nécessaire, espelho compacto",
        "Pessoas interessadas em organização, rotina e estética sem cosméticos",
        "Parcial",
        "Evitar cosméticos, dermocosméticos, clareadores e promessas reguladas pela Anvisa.",
    ],
]


PRODUCTS = [
    [
        "Casa, organização e utilidades práticas",
        "Organizador de cabos",
        "Reduz bagunça e risco de fios soltos.",
        "Residências, home office e escritórios",
        14.90,
        8.90,
        49.90,
        8.00,
        7.50,
        0.28,
        "Baixo",
        "Dropi",
        "Vídeo de instalação em menos de 10 segundos",
        "Produto simples, visual e fácil de explicar.",
    ],
    [
        "Casa, organização e utilidades práticas",
        "Suporte adesivo para tomada",
        "Organiza tomadas, carregadores e pequenos aparelhos.",
        "Famílias e pessoas com pouco espaço",
        11.50,
        8.50,
        39.90,
        7.00,
        6.50,
        0.30,
        "Baixo",
        "AliExpress + DSers",
        "Antes e depois na cozinha ou escritório",
        "Validar cola e acabamento com amostra.",
    ],
    [
        "Pet utilitário",
        "Escova removedora de pelos pet",
        "Remove pelos de sofá, roupa e tapete.",
        "Tutores de cães e gatos",
        18.50,
        9.50,
        59.90,
        8.50,
        8.00,
        0.27,
        "Baixo",
        "Dropi",
        "Demonstração real em tecido escuro",
        "Apelo visual forte; não fazer promessa veterinária.",
    ],
    [
        "Pet utilitário",
        "Luva removedora de pelos",
        "Ajuda a tirar pelos soltos durante carinho e escovação.",
        "Tutores de pets que soltam pelo",
        13.90,
        8.90,
        49.90,
        7.50,
        7.00,
        0.28,
        "Baixo",
        "AliExpress + DSers",
        "Uso real com pet calmo",
        "Evitar linguagem de tratamento veterinário.",
    ],
    [
        "Casa, organização e utilidades práticas",
        "Organizador de geladeira",
        "Melhora visualização dos alimentos e reduz desperdício.",
        "Famílias e pessoas que preparam refeições",
        21.50,
        9.50,
        69.90,
        9.50,
        9.00,
        0.26,
        "Médio",
        "DropNacional",
        "Transformação visual da geladeira",
        "Checar dimensão, material e embalagem.",
    ],
    [
        "Cozinha prática",
        "Cortador/descascador prático",
        "Economiza tempo no preparo de alimentos.",
        "Quem cozinha em casa",
        16.80,
        9.20,
        54.90,
        8.50,
        8.00,
        0.27,
        "Baixo",
        "Dropi",
        "Uso real com alimento comum",
        "Evitar prometer desempenho exagerado.",
    ],
    [
        "Automotivo leve",
        "Suporte veicular para celular",
        "Facilita visualização do celular com mais organização.",
        "Motoristas e motoristas de aplicativo",
        17.40,
        10.00,
        59.90,
        8.50,
        8.50,
        0.26,
        "Baixo",
        "AliExpress + DSers",
        "Instalação no painel com tomada curta",
        "Checar compatibilidade e fixação.",
    ],
    [
        "Automotivo leve",
        "Lixeira dobrável para carro",
        "Mantém o carro limpo e ocupa pouco espaço.",
        "Motoristas, famílias e viagens",
        19.20,
        9.80,
        64.90,
        8.50,
        8.50,
        0.26,
        "Baixo",
        "Dropi",
        "Antes/depois do interior do carro",
        "Boa demonstração para Reels.",
    ],
    [
        "Automotivo leve",
        "Organizador de banco de carro",
        "Evita objetos soltos no banco traseiro.",
        "Famílias, motoristas e pessoas que viajam",
        22.30,
        10.50,
        69.90,
        9.00,
        9.00,
        0.25,
        "Médio",
        "AliExpress + DSers",
        "Demonstração em carro real",
        "Validar costura e tamanho.",
    ],
    [
        "Pet utilitário",
        "Tapete higiênico lavável para pet",
        "Ajuda na rotina de higiene e reduz descartáveis.",
        "Tutores de cães em apartamento",
        29.50,
        12.00,
        89.90,
        10.00,
        10.00,
        0.25,
        "Médio",
        "DropNacional",
        "Rotina de limpeza sem promessa veterinária",
        "Amostra é essencial para validar absorção.",
    ],
    [
        "Pet utilitário",
        "Bebedouro portátil para pet",
        "Facilita hidratação em passeios e viagens.",
        "Tutores que passeiam com pets",
        26.00,
        11.50,
        79.90,
        9.50,
        9.50,
        0.25,
        "Médio",
        "Dropi",
        "Uso real em passeio",
        "Validar vedação e material.",
    ],
    [
        "Cozinha prática",
        "Porta-temperos organizador",
        "Organiza temperos e libera espaço.",
        "Famílias e cozinhas pequenas",
        20.00,
        9.20,
        59.90,
        8.50,
        8.00,
        0.26,
        "Baixo",
        "DropNacional",
        "Antes/depois do armário",
        "Boa aceitação em conteúdo de organização.",
    ],
    [
        "Casa, organização e utilidades práticas",
        "Escova de limpeza multifuncional",
        "Facilita limpeza de cantos e superfícies.",
        "Famílias e pessoas que limpam a casa",
        18.30,
        9.00,
        54.90,
        8.00,
        7.50,
        0.27,
        "Baixo",
        "AliExpress + DSers",
        "Demonstração rápida de sujeira real",
        "Produto barato, mas precisa de bom criativo.",
    ],
    [
        "Casa, organização e utilidades práticas",
        "Vedador de portas",
        "Ajuda a reduzir entrada de poeira, vento e ruído.",
        "Residências e apartamentos",
        15.80,
        9.50,
        49.90,
        8.00,
        7.00,
        0.28,
        "Baixo",
        "Dropi",
        "Problema cotidiano com antes/depois",
        "Não prometer isolamento absoluto.",
    ],
    [
        "Casa, organização e utilidades práticas",
        "Luminária sensor de presença simples",
        "Ilumina passagens e armários sem instalação complexa.",
        "Residências e pequenos espaços",
        24.90,
        11.50,
        79.90,
        9.50,
        9.50,
        0.25,
        "Médio",
        "Dropi",
        "Ambiente escuro antes/depois",
        "Evitar modelos elétricos complexos ou bateria ruim.",
    ],
    [
        "Casa, organização e utilidades práticas",
        "Saco organizador a vácuo",
        "Economiza espaço em armários e malas.",
        "Famílias, apartamentos e viagens",
        19.00,
        9.60,
        64.90,
        8.50,
        8.50,
        0.26,
        "Baixo",
        "Dropi",
        "Antes/depois com cobertor ou roupa",
        "Checar vedação e tamanho real.",
    ],
    [
        "Cozinha prática",
        "Escorredor dobrável",
        "Organiza a pia e ocupa menos espaço.",
        "Cozinhas pequenas",
        16.20,
        9.50,
        54.90,
        8.00,
        7.50,
        0.27,
        "Baixo",
        "AliExpress + DSers",
        "Demonstração na pia",
        "Validar material e peso suportado.",
    ],
    [
        "Casa, organização e utilidades práticas",
        "Bolsa organizadora de mala",
        "Facilita separar roupas e acessórios em viagens.",
        "Pessoas que viajam e famílias",
        21.20,
        10.50,
        69.90,
        9.00,
        8.50,
        0.26,
        "Baixo",
        "DropNacional",
        "Mala bagunçada versus organizada",
        "Boa pauta para conteúdo orgânico.",
    ],
    [
        "Beleza não regulada",
        "Espelho ou organizador de maquiagem sem cosmético",
        "Organiza itens de beleza sem envolver produto regulado.",
        "Pessoas com rotina de maquiagem",
        18.00,
        9.50,
        59.90,
        8.50,
        8.00,
        0.27,
        "Baixo",
        "Dropi",
        "Organização de bancada",
        "Não vender como cosmético ou item terapêutico.",
    ],
    [
        "Fitness leve e bem-estar sem promessa médica",
        "Garrafa ou shaker sem promessa de saúde",
        "Facilita transporte de bebida e rotina de treino.",
        "Pessoas que treinam ou trabalham fora",
        19.50,
        9.50,
        64.90,
        8.50,
        8.50,
        0.26,
        "Baixo",
        "Dropi",
        "Uso real na rotina",
        "Não associar a emagrecimento ou benefício médico.",
    ],
    [
        "Casa, organização e utilidades práticas",
        "Suporte de parede para vassouras",
        "Organiza área de serviço e evita itens caídos.",
        "Casas, apartamentos e lavanderias",
        17.90,
        9.90,
        59.90,
        8.50,
        8.00,
        0.27,
        "Baixo",
        "Fornecedor nacional direto",
        "Antes/depois da lavanderia",
        "Checar adesivo, parafusos e resistência.",
    ],
    [
        "Cozinha prática",
        "Tampa de silicone multiuso",
        "Ajuda a conservar alimentos e reduzir sujeira.",
        "Famílias e pessoas que cozinham",
        13.80,
        8.90,
        44.90,
        7.50,
        6.50,
        0.28,
        "Baixo",
        "DropNacional",
        "Teste real com potes de tamanhos diferentes",
        "Validar material e embalagem individual.",
    ],
]


FORNECEDORES = [
    [
        "Dropi",
        "Plataforma brasileira",
        "Centralizar operação de dropshipping no Brasil",
        "Integrações e operação mais simples para iniciantes",
        "Plano, catálogo e fornecedores disponíveis mudam com o tempo",
        "3 a 10 dias",
        "Nuvemshop, Shopify, WooCommerce, Yampi, AliExpress, CJdropshipping e SourcinBox",
        "Variável",
        "Alta",
        "Sim",
        "Validar plano atual, catálogo, logística, prazo e reputação",
        "validar no site oficial",
    ],
    [
        "DropNacional",
        "Rede de fornecedores nacionais",
        "Reduzir prazo e evitar longa importação",
        "Prazo menor e mais controle sobre entrega",
        "Catálogo pode ser menor e margem pode apertar",
        "1 a 7 dias",
        "Nuvemshop, Shopify, WooCommerce e operação manual",
        "Variável",
        "Alta",
        "Sim",
        "Verificar catálogo, margem, reputação e política de troca",
        "validar no site oficial",
    ],
    [
        "AliExpress + DSers",
        "Marketplace internacional + app de pedidos",
        "Testar produtos com grande variedade",
        "Muito conhecido, catálogo enorme e fácil de encontrar tendências",
        "Prazo pode ser alto, qualidade varia e exige análise rigorosa",
        "10 a 25 dias",
        "Shopify, WooCommerce e Nuvemshop conforme integração disponível",
        "Gratuito/variável",
        "Média",
        "Sim com cautela",
        "Analisar avaliações, pedidos, fotos reais, prazo para o Brasil e rastreio",
        "validar no site oficial",
    ],
    [
        "CJdropshipping",
        "Sourcing e fulfillment internacional",
        "Escalar produtos já validados",
        "Sourcing, fulfillment e variedade de produtos",
        "Prazo, frete e disponibilidade de armazém variam",
        "7 a 20 dias",
        "Shopify, WooCommerce e outras integrações",
        "Variável",
        "Média",
        "Parcial",
        "Validar prazo, frete, armazéns disponíveis e custo final para o Brasil",
        "validar no site oficial",
    ],
    [
        "SourcinBox",
        "Sourcing/fulfillment",
        "Encontrar fornecedores e melhorar logística",
        "Pode ajudar no sourcing e na operação quando há volume",
        "Menos útil antes da validação do produto",
        "7 a 20 dias",
        "Shopify, WooCommerce e operação conforme disponibilidade",
        "Variável",
        "Média",
        "Não no primeiro teste",
        "Usar após validar produto, margem e demanda",
        "validar no site oficial",
    ],
    [
        "Zendrop",
        "Plataforma internacional",
        "Alternativa estruturada para dropshipping",
        "Boa estrutura operacional e automações",
        "Pode ter custo em dólar e logística menos vantajosa para Brasil",
        "7 a 20 dias",
        "Shopify, WooCommerce e integrações disponíveis",
        "Variável/dólar",
        "Média",
        "Parcial",
        "Comparar custo em reais, prazo e suporte ao Brasil",
        "validar no site oficial",
    ],
    [
        "Spocket",
        "Fornecedores EUA/Europa",
        "Alternativa internacional",
        "Catálogo curado e bons fornecedores em alguns mercados",
        "Pode não ser ideal para Brasil no começo",
        "7 a 25 dias",
        "Shopify, WooCommerce e integrações disponíveis",
        "Variável/dólar",
        "Média",
        "Não no começo",
        "Verificar prazo, frete e custo final para clientes no Brasil",
        "validar no site oficial",
    ],
    [
        "Fornecedor nacional direto",
        "Fabricante/distribuidor",
        "Ter prazo menor e mais controle",
        "Negociação direta, melhor prazo e possível margem maior",
        "Exige prospecção, negociação e controle manual",
        "1 a 7 dias",
        "Manual, WhatsApp, planilha ou ERP simples",
        "Variável",
        "Alta se validado",
        "Sim",
        "Confirmar CNPJ, reputação, nota fiscal, prazo, política de troca e estoque",
        "validar no site oficial",
    ],
    [
        "Mercado Livre atacado",
        "Marketplace B2B/atacado",
        "Testar produtos com envio mais rápido",
        "Acesso fácil a vendedores nacionais e entrega previsível",
        "Margem e reputação variam muito",
        "2 a 7 dias",
        "Manual",
        "Variável",
        "Média",
        "Parcial",
        "Verificar reputação, nota fiscal, estoque, margem e autorização de revenda",
        "validar no site oficial",
    ],
    [
        "Shopee fornecedores/atacado",
        "Marketplace com vendedores e atacadistas",
        "Buscar itens de baixo investimento",
        "Catálogo amplo e preços competitivos",
        "Qualidade, prazo e reputação precisam ser verificados com rigor",
        "5 a 15 dias",
        "Manual",
        "Variável",
        "Média",
        "Parcial",
        "Comprar amostra, verificar avaliações, reputação e documentação",
        "validar no site oficial",
    ],
]


def default_output_dir() -> Path:
    preferred = Path("/mnt/data")
    try:
        preferred.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(dir=preferred, delete=True):
            pass
        return preferred
    except OSError:
        fallback = Path.cwd() / "outputs"
        fallback.mkdir(parents=True, exist_ok=True)
        return fallback


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Gera uma planilha Excel profissional para planejamento de dropshipping no Brasil."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Caminho do arquivo .xlsx gerado.",
    )
    parser.add_argument(
        "--preview",
        type=Path,
        default=None,
        help="Caminho da imagem PNG de prévia do dashboard.",
    )
    parser.add_argument(
        "--skip-preview",
        action="store_true",
        help="Não gera a imagem PNG de prévia.",
    )
    args = parser.parse_args()
    if args.output is None or args.preview is None:
        output_dir = default_output_dir()
        args.output = args.output or output_dir / WORKBOOK_NAME
        args.preview = args.preview or output_dir / PREVIEW_NAME
    return args


def safe_table_name(sheet_title: str, suffix: str = "Tabela") -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in sheet_title)
    return f"{suffix}_{cleaned}".strip("_")[:240]


def set_doc_properties(wb: Workbook) -> None:
    wb.properties.creator = "Codex"
    wb.properties.title = "Planejamento Dropshipping Brasil"
    wb.properties.subject = "Planejamento realista para dropshipping com baixo risco inicial"
    wb.properties.description = (
        "Planilha profissional com nichos, produtos, fornecedores, custos, "
        "precificação, marketing, riscos, legal, cronograma e projeção financeira."
    )
    wb.properties.keywords = "dropshipping, Brasil, planejamento, precificação, riscos"
    wb.properties.category = "Planejamento"


def new_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    set_doc_properties(wb)
    return wb


def style_title(cell, size: int = 16) -> None:
    cell.font = Font(size=size, bold=True, color=PALETTE.navy)
    cell.alignment = Alignment(vertical="center", wrap_text=True)


def style_note(cell) -> None:
    cell.fill = NOTE_FILL
    cell.font = Font(italic=True, color=PALETTE.gray)
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = THIN_BORDER


def write_title(ws: Worksheet, title: str, subtitle: str | None = None, end_col: int = 8) -> None:
    ws["A1"] = title
    style_title(ws["A1"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.row_dimensions[1].height = 30
    if subtitle:
        ws["A2"] = subtitle
        style_note(ws["A2"])
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
        ws.row_dimensions[2].height = 36


def write_table(
    ws: Worksheet,
    headers: Sequence[str],
    rows: Sequence[Sequence[object]],
    *,
    start_row: int = 1,
    start_col: int = 1,
    table_name: str | None = None,
    table_style: str = "TableStyleMedium2",
) -> tuple[int, int]:
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + col_offset, value=header)
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color=PALETTE.white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    for row_offset, row_values in enumerate(rows, start=1):
        for col_offset, value in enumerate(row_values):
            cell = ws.cell(row=start_row + row_offset, column=start_col + col_offset, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER

    end_row = start_row + len(rows)
    end_col = start_col + len(headers) - 1
    if rows:
        ref = (
            f"{get_column_letter(start_col)}{start_row}:"
            f"{get_column_letter(end_col)}{end_row}"
        )
        display_name = table_name or safe_table_name(ws.title)
        table = Table(displayName=display_name, ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
        ws.auto_filter.ref = ref
    return end_row, end_col


def set_column_widths(
    ws: Worksheet,
    *,
    min_width: int = 10,
    max_width: int = 42,
    extra: int = 2,
    widths: dict[str, float] | None = None,
) -> None:
    widths = widths or {}
    for column_cells in ws.columns:
        col_letter = get_column_letter(column_cells[0].column)
        if col_letter in widths:
            ws.column_dimensions[col_letter].width = widths[col_letter]
            continue
        max_len = 0
        for cell in column_cells:
            value = cell.value
            if value is None:
                continue
            max_len = max(max_len, min(len(str(value)), max_width))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + extra))


def apply_sheet_finish(ws: Worksheet, freeze: str = "A2") -> None:
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.print_title_rows = "1:1"
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical="top",
                wrap_text=True,
            )
            if cell.border == Border():
                cell.border = THIN_BORDER
            if isinstance(cell.value, str) and cell.value.startswith("="):
                cell.font = Font(color=PALETTE.blue, bold=cell.font.bold)
    set_column_widths(ws)


def add_list_validation(
    ws: Worksheet,
    cell_range: str,
    values: Sequence[str],
    *,
    allow_blank: bool = True,
    prompt: str | None = None,
) -> None:
    validation = DataValidation(
        type="list",
        formula1=f'"{",".join(values)}"',
        allow_blank=allow_blank,
    )
    validation.error = "Escolha uma opção da lista."
    validation.errorTitle = "Valor inválido"
    if prompt:
        validation.prompt = prompt
        validation.promptTitle = "Opções"
    ws.add_data_validation(validation)
    validation.add(cell_range)


def apply_number_formats(
    ws: Worksheet,
    *,
    currency_cols: Iterable[str] = (),
    percent_cols: Iterable[str] = (),
    integer_cols: Iterable[str] = (),
    start_row: int = 2,
) -> None:
    for col in currency_cols:
        for row in range(start_row, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = CURRENCY_FORMAT
    for col in percent_cols:
        for row in range(start_row, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = PERCENT_FORMAT
    for col in integer_cols:
        for row in range(start_row, ws.max_row + 1):
            ws[f"{col}{row}"].number_format = INTEGER_FORMAT


def add_status_validation(ws: Worksheet, cell_range: str) -> None:
    add_list_validation(
        ws,
        cell_range,
        ["Pendente", "Em andamento", "Feito", "Bloqueado"],
        allow_blank=False,
        prompt="Escolha: Pendente, Em andamento, Feito ou Bloqueado.",
    )
    first_row = int("".join(ch for ch in cell_range.split(":")[0] if ch.isdigit()))
    first_col = "".join(ch for ch in cell_range.split(":")[0] if ch.isalpha())
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'=${first_col}{first_row}="Feito"'],
            fill=GOOD_FILL,
            font=Font(color=PALETTE.green, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f'=${first_col}{first_row}="Bloqueado"'],
            fill=BAD_FILL,
            font=Font(color=PALETTE.red, bold=True),
        ),
    )


def add_yes_no_validation(ws: Worksheet, cell_range: str) -> None:
    add_list_validation(ws, cell_range, ["Sim", "Não", "Parcial"])


def add_priority_validation(ws: Worksheet, cell_range: str) -> None:
    add_list_validation(ws, cell_range, ["Baixa", "Média", "Alta"])


def add_risk_validation(ws: Worksheet, cell_range: str) -> None:
    add_list_validation(ws, cell_range, ["Baixo", "Médio", "Alto"])
    add_text_alerts(
        ws,
        cell_range,
        good_values=["Baixo"],
        warning_values=["Médio"],
        bad_values=["Alto"],
    )


def add_flexible_yes_no_validation(ws: Worksheet, cell_range: str) -> None:
    add_list_validation(
        ws,
        cell_range,
        ["Sim", "Sim com cautela", "Parcial", "Não", "Não no começo", "Não no primeiro teste", "Trial", "Validar"],
    )


def add_text_alerts(
    ws: Worksheet,
    cell_range: str,
    *,
    good_values: Sequence[str] = (),
    warning_values: Sequence[str] = (),
    bad_values: Sequence[str] = (),
) -> None:
    first = cell_range.split(":")[0]
    first_row = int("".join(ch for ch in first if ch.isdigit()))
    first_col = "".join(ch for ch in first if ch.isalpha())

    def formula(values: Sequence[str]) -> str:
        if len(values) == 1:
            return f'={first_col}{first_row}="{values[0]}"'
        clauses = ",".join(f'{first_col}{first_row}="{value}"' for value in values)
        return f"=OR({clauses})"

    if good_values:
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[formula(good_values)],
                fill=GOOD_FILL,
                font=Font(color=PALETTE.green, bold=True),
            ),
        )
    if warning_values:
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[formula(warning_values)],
                fill=WARN_FILL,
                font=Font(color="7A4D00", bold=True),
            ),
        )
    if bad_values:
        ws.conditional_formatting.add(
            cell_range,
            FormulaRule(
                formula=[formula(bad_values)],
                fill=BAD_FILL,
                font=Font(color=PALETTE.red, bold=True),
            ),
        )


def add_numeric_alerts(ws: Worksheet, cell_range: str) -> None:
    first = cell_range.split(":")[0]
    first_row = int("".join(ch for ch in first if ch.isdigit()))
    first_col = "".join(ch for ch in first if ch.isalpha())
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f"={first_col}{first_row}>=0.30"],
            fill=GOOD_FILL,
            font=Font(color=PALETTE.green, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f"=AND({first_col}{first_row}>=0.20,{first_col}{first_row}<0.30)"],
            fill=WARN_FILL,
            font=Font(color="7A4D00", bold=True),
        ),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(
            formula=[f"={first_col}{first_row}<0.20"],
            fill=BAD_FILL,
            font=Font(color=PALETTE.red, bold=True),
        ),
    )


def add_decimal_validation(
    ws: Worksheet,
    cell_range: str,
    *,
    minimum: float,
    maximum: float,
    error: str,
) -> None:
    validation = DataValidation(
        type="decimal",
        operator="between",
        formula1=str(minimum),
        formula2=str(maximum),
        allow_blank=False,
    )
    validation.error = error
    validation.errorTitle = "Número inválido"
    ws.add_data_validation(validation)
    validation.add(cell_range)


def add_kpi_card(ws: Worksheet, start_cell: str, label: str, value: str, fill_color: str) -> None:
    row = ws[start_cell].row
    col = ws[start_cell].column
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 2)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 2)
    label_cell = ws.cell(row=row, column=col, value=label)
    value_cell = ws.cell(row=row + 1, column=col, value=value)
    for r in range(row, row + 3):
        for c in range(col, col + 3):
            cell = ws.cell(row=r, column=c)
            cell.fill = PatternFill("solid", fgColor=fill_color)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    label_cell.font = Font(size=10, bold=True, color=PALETTE.white)
    value_cell.font = Font(size=12, bold=True, color=PALETTE.white)
    ws.row_dimensions[row].height = 22
    ws.row_dimensions[row + 1].height = 28
    ws.row_dimensions[row + 2].height = 28


def build_dashboard(wb: Workbook) -> None:
    ws = wb.create_sheet("00 Escolha_Principal")
    ws.sheet_properties.tabColor = PALETTE.blue
    write_title(
        ws,
        "Planejamento de Dropshipping no Brasil",
        "Decisão executiva para iniciar com menor risco: foco em utilidade clara, margem realista, fornecedores validados e testes pequenos.",
        end_col=16,
    )

    add_kpi_card(ws, "A4", "Nicho recomendado", "Casa, organização e utilidades práticas", PALETTE.blue)
    add_kpi_card(ws, "D4", "Capital inicial", "R$ 400 a R$ 1.000", PALETTE.green)
    add_kpi_card(ws, "G4", "Fornecedor inicial", "Dropi ou nacional direto", PALETTE.purple)
    add_kpi_card(ws, "J4", "Canal orgânico", "TikTok + Reels", PALETTE.amber)
    add_kpi_card(ws, "M4", "Canal pago", "Meta Ads + TikTok Ads", PALETTE.navy)

    rows = [
        ["Produto inicial mais seguro", "Organizador de cabos ou suporte adesivo para tomada"],
        ["Principais riscos", "Prazo longo, fornecedor ruim, margem apertada, reembolso e criativo fraco"],
        [
            "Estratégia inicial",
            "Começar com 3 a 5 produtos simples, comprar amostra, criar página limpa, validar criativos orgânicos e só então testar anúncios com orçamento controlado.",
        ],
        [
            "Caminho de menor risco",
            "Priorizar fornecedor nacional ou plataforma com rastreio, política clara, checkout simples, atendimento em até 24h e reserva para reembolso.",
        ],
        [
            "Regra de decisão",
            "Escalar apenas quando houver margem líquida positiva, entrega previsível, criativo validado e reclamações sob controle.",
        ],
    ]
    write_table(
        ws,
        ["Resumo executivo", "Decisão recomendada"],
        rows,
        start_row=9,
        table_name="Tabela_Dashboard_Resumo",
        table_style="TableStyleMedium9",
    )

    ranking_rows = [[row[1], row[2], row[10]] for row in NICHOS]
    write_table(
        ws,
        ["Nicho", "Nota", "Recomendado"],
        ranking_rows,
        start_row=17,
        table_name="Tabela_Dashboard_Nichos",
        table_style="TableStyleMedium4",
    )

    cost_rows = [
        ["Econômico", "='06 Custos_Iniciais'!C17"],
        ["Intermediário", "='06 Custos_Iniciais'!D17"],
        ["Profissional", "='06 Custos_Iniciais'!E17"],
    ]
    write_table(
        ws,
        ["Cenário", "Investimento estimado"],
        cost_rows,
        start_row=17,
        start_col=5,
        table_name="Tabela_Dashboard_Custos",
        table_style="TableStyleMedium7",
    )
    for row in range(18, 21):
        ws[f"F{row}"].number_format = CURRENCY_FORMAT

    projection_rows = [
        ["Ruim", "='15 Financeiro_3_Meses'!N7"],
        ["Médio", "='15 Financeiro_3_Meses'!N10"],
        ["Bom", "='15 Financeiro_3_Meses'!N13"],
    ]
    write_table(
        ws,
        ["Cenário", "Lucro estimado no mês 3"],
        projection_rows,
        start_row=17,
        start_col=8,
        table_name="Tabela_Dashboard_Lucro",
        table_style="TableStyleMedium11",
    )
    for row in range(18, 21):
        ws[f"I{row}"].number_format = CURRENCY_FORMAT

    avoid_rows = [
        ["Suplementos e produtos médicos", "Risco regulatório, promessa sensível e bloqueios de anúncio"],
        ["Cosméticos com promessa forte", "Risco Anvisa, reembolso e publicidade enganosa"],
        ["Eletrônicos complexos ou com bateria ruim", "Defeito, devolução e suporte difícil"],
        ["Moda com numeração", "Trocas frequentes e variação de tamanho"],
        ["Produtos falsificados, réplicas e marcas sem autorização", "Risco legal e bloqueio de conta"],
        ["Produtos infantis com risco de segurança", "Alto risco operacional e reputacional"],
    ]
    write_table(
        ws,
        ["Evitar no começo", "Motivo"],
        avoid_rows,
        start_row=17,
        start_col=11,
        table_name="Tabela_Dashboard_Evitar",
        table_style="TableStyleMedium3",
    )

    bar = BarChart()
    bar.type = "bar"
    bar.style = 10
    bar.title = "Ranking de nichos por nota"
    bar.y_axis.title = "Nicho"
    bar.x_axis.title = "Nota"
    bar.height = 7
    bar.width = 10
    bar.add_data(Reference(ws, min_row=17, max_row=23, min_col=2), titles_from_data=True)
    bar.set_categories(Reference(ws, min_row=18, max_row=23, min_col=1))
    ws.add_chart(bar, "A26")

    costs = BarChart()
    costs.type = "col"
    costs.style = 12
    costs.title = "Custos por cenário"
    costs.y_axis.title = "R$"
    costs.height = 7
    costs.width = 9
    costs.add_data(Reference(ws, min_row=17, max_row=20, min_col=6), titles_from_data=True)
    costs.set_categories(Reference(ws, min_row=18, max_row=20, min_col=5))
    ws.add_chart(costs, "H26")

    profit = LineChart()
    profit.style = 13
    profit.title = "Lucro estimado no mês 3"
    profit.y_axis.title = "R$"
    profit.height = 7
    profit.width = 9
    profit.add_data(Reference(ws, min_row=17, max_row=20, min_col=9), titles_from_data=True)
    profit.set_categories(Reference(ws, min_row=18, max_row=20, min_col=8))
    ws.add_chart(profit, "M26")

    ws["A42"] = (
        "Observação: a planilha usa estimativas realistas e não promete lucro. "
        "Dados de plataformas, preços, integrações, regras e tributos devem ser validados em fontes oficiais."
    )
    ws.merge_cells("A42:P43")
    style_note(ws["A42"])
    for row in range(42, 44):
        for col in range(1, 17):
            ws.cell(row=row, column=col).fill = NOTE_FILL
            ws.cell(row=row, column=col).border = THIN_BORDER

    set_column_widths(
        ws,
        widths={
            "A": 28,
            "B": 13,
            "C": 14,
            "D": 18,
            "E": 20,
            "F": 16,
            "G": 16,
            "H": 18,
            "I": 18,
            "J": 14,
            "K": 30,
            "L": 40,
            "M": 14,
            "N": 14,
            "O": 14,
            "P": 14,
        },
    )
    ws.freeze_panes = "A9"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 85
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def build_overview(wb: Workbook) -> None:
    ws = wb.create_sheet("01 Visao_Geral")
    ws.sheet_properties.tabColor = PALETTE.green
    rows = [
        ["Como funciona", "O cliente compra na loja, você compra do fornecedor e o fornecedor envia diretamente ao cliente."],
        ["Etapas da venda", "Pesquisa, validação, loja, tráfego, pedido, compra no fornecedor, envio, rastreamento, atendimento e pós-venda."],
        ["Como o lucro é formado", "Lucro é preço de venda menos produto, frete, taxas, anúncios, impostos, reembolsos, ferramentas e suporte."],
        ["Principal cuidado", "Não anunciar produto sem validar amostra, prazo, margem e política de troca."],
        ["O que não fazer", "Não vender produto proibido, falsificado, perigoso, com promessa médica ou sem margem realista."],
        ["Expectativa realista", "Pode gerar lucro, mas exige teste, atendimento, controle financeiro e tolerância a erro controlado."],
        ["Risco inicial", "Atraso, fornecedor ruim, reembolso, chargeback e conta de anúncios bloqueada podem inviabilizar a operação."],
        ["Métrica de saúde", "Margem líquida positiva, reclamações baixas, entrega previsível e criativo com CPA aceitável."],
        ["Caminho seguro", "Começar pequeno, documentar tudo, medir cada etapa e reinvestir apenas parte do lucro."],
    ]
    write_table(ws, ["Tema", "Resumo prático"], rows, table_name="Tabela_Visao_Geral")
    apply_sheet_finish(ws)


def build_business_model(wb: Workbook) -> None:
    ws = wb.create_sheet("02 Modelo_Negocio")
    ws.sheet_properties.tabColor = "17BECF"
    rows = [
        ["Pesquisa de nicho", "Mapear dores, intenção de compra e concorrência.", "Google Trends, TikTok, Instagram, Mercado Livre", "Escolher nicho só por gosto pessoal.", "Comparar demanda, margem e risco."],
        ["Escolha do produto", "Priorizar utilidade clara, ticket acessível e baixo risco de defeito.", "Dropi, DSers, DropNacional, marketplaces", "Comprar produto apenas porque viralizou.", "Avaliar qualidade, prazo, avaliações e margem."],
        ["Validação", "Confirmar amostra, criativo, fornecedor, prazo e preço.", "Planilha, pedido teste, vídeos orgânicos", "Anunciar sem testar.", "Comprar amostra e medir resposta orgânica."],
        ["Fornecedor", "Verificar reputação, estoque, rastreio e política de troca.", "Dropi, DropNacional, fornecedor direto", "Depender de fornecedor sem contato.", "Ter fornecedor alternativo e SLA mínimo."],
        ["Criação da loja", "Montar páginas claras, políticas, checkout simples e prova social honesta.", "Nuvemshop, Shopify, Yampi, WooCommerce", "Loja com aparência improvisada.", "Usar layout limpo, imagens reais e texto objetivo."],
        ["Anúncio", "Testar criativos com orçamento pequeno e metas claras.", "Meta Ads, TikTok Ads", "Escalar sem dados.", "Definir CPA máximo e regra de pausa."],
        ["Venda", "Receber pedido com prazo e comunicação transparentes.", "Loja, gateway, WhatsApp Business", "Prometer prazo irreal.", "Informar prazo conservador e rastreio."],
        ["Compra no fornecedor", "Comprar corretamente após a venda e registrar evidências.", "Fornecedor, planilha, automação", "Comprar item errado ou sem rastreio.", "Padronizar SKU e checklist de pedido."],
        ["Envio", "Acompanhar despacho, rastreio e exceções.", "Melhor Envio, rastreio do fornecedor", "Não monitorar pedido.", "Rotina diária de rastreio."],
        ["Rastreamento", "Avisar cliente sobre status e atraso antes da reclamação.", "E-mail, WhatsApp, página de rastreio", "Cliente sem informação.", "Enviar atualizações claras."],
        ["Atendimento", "Responder rápido, com registro e linguagem profissional.", "WhatsApp Business, e-mail, helpdesk", "Responder tarde ou de forma genérica.", "Criar respostas rápidas e SLA de 24h."],
        ["Troca/devolução", "Aplicar política clara e compatível com as regras brasileiras.", "Política da loja, contador/fonte oficial", "Improvisar regra depois do problema.", "Publicar política antes de vender."],
        ["Reinvestimento", "Separar reserva, impostos, reembolso e verba de teste.", "Planilha financeira", "Gastar todo o lucro aparente.", "Definir percentual de reinvestimento e reserva."],
    ]
    write_table(
        ws,
        ["Etapa", "O que fazer", "Ferramenta sugerida", "Erro comum", "Como evitar"],
        rows,
        table_name="Tabela_Modelo_Negocio",
    )
    apply_sheet_finish(ws)


def build_niches(wb: Workbook) -> None:
    ws = wb.create_sheet("03 Nichos")
    ws.sheet_properties.tabColor = PALETTE.purple
    headers = [
        "Ranking",
        "Nicho",
        "Nota de 0 a 10",
        "Potencial de venda",
        "Concorrência",
        "Ticket médio estimado",
        "Facilidade de anunciar",
        "Risco operacional",
        "Produtos indicados",
        "Perfil do cliente",
        "Recomendado para iniciante",
        "Justificativa",
    ]
    write_table(ws, headers, NICHOS, table_name="Tabela_Nichos", table_style="TableStyleMedium4")
    add_yes_no_validation(ws, f"K2:K{ws.max_row}")
    for row in range(2, ws.max_row + 1):
        ws[f"C{row}"].number_format = "0.0"
    apply_sheet_finish(ws)


def build_products(wb: Workbook) -> None:
    ws = wb.create_sheet("04 Produtos")
    ws.sheet_properties.tabColor = "FF7F0E"
    headers = [
        "Nicho",
        "Produto",
        "Problema que resolve",
        "Público-alvo",
        "Custo estimado",
        "Frete estimado",
        "Preço de venda sugerido",
        "Taxas estimadas",
        "CPA máximo aceitável",
        "Margem estimada",
        "Risco",
        "Fornecedor sugerido",
        "Tipo de criativo ideal",
        "Observações",
    ]
    write_table(ws, headers, PRODUCTS, table_name="Tabela_Produtos", table_style="TableStyleMedium6")
    for row in range(2, ws.max_row + 1):
        ws[f"J{row}"] = f'=IF(G{row}=0,0,(G{row}-E{row}-F{row}-H{row}-I{row})/G{row})'
    apply_number_formats(
        ws,
        currency_cols=["E", "F", "G", "H", "I"],
        percent_cols=["J"],
        start_row=2,
    )
    add_numeric_alerts(ws, f"J2:J{ws.max_row}")
    add_risk_validation(ws, f"K2:K{ws.max_row}")
    apply_sheet_finish(ws)


def build_suppliers(wb: Workbook) -> None:
    ws = wb.create_sheet("05 Fornecedores")
    ws.sheet_properties.tabColor = PALETTE.red
    headers = [
        "Fornecedor/Plataforma",
        "Tipo",
        "Melhor uso",
        "Vantagens",
        "Desvantagens",
        "Prazo estimado",
        "Integrações",
        "Custo aproximado",
        "Nível de confiança",
        "Indicado para iniciante",
        "Cuidados antes de usar",
        "Link oficial ou campo para validação",
    ]
    write_table(ws, headers, FORNECEDORES, table_name="Tabela_Fornecedores", table_style="TableStyleMedium3")
    add_text_alerts(
        ws,
        f"I2:I{ws.max_row}",
        good_values=["Alta", "Alta se validado"],
        warning_values=["Média"],
    )
    add_flexible_yes_no_validation(ws, f"J2:J{ws.max_row}")
    add_text_alerts(
        ws,
        f"J2:J{ws.max_row}",
        good_values=["Sim"],
        warning_values=["Sim com cautela", "Parcial"],
        bad_values=["Não", "Não no começo", "Não no primeiro teste"],
    )
    apply_sheet_finish(ws)


def build_initial_costs(wb: Workbook) -> None:
    ws = wb.create_sheet("06 Custos_Iniciais")
    ws.sheet_properties.tabColor = "8C564B"
    rows = [
        ["Loja", "Plataforma de loja", 29, 79, 299, "Plano mensal inicial ou teste pago."],
        ["Domínio", "Domínio", 40, 70, 120, "Nome próprio aumenta confiança."],
        ["Design", "Tema", 0, 150, 500, "Tema gratuito, pago ou customização."],
        ["Apps", "Apps e integrações", 0, 120, 600, "Pagamento, avaliações, logística e recuperação de carrinho."],
        ["Integração", "Dropi ou ferramenta de integração", 0, 99, 299, "Usar se realmente reduzir trabalho operacional."],
        ["Marketing", "Anúncios", 150, 800, 3000, "Testes controlados; não escalar sem margem."],
        ["Validação", "Compra de amostras", 80, 250, 600, "Obrigatório para produto principal."],
        ["Conteúdo", "Edição de vídeo", 0, 250, 1000, "CapCut próprio, freelancer ou pacote de criativos."],
        ["Conteúdo", "Criativos", 50, 200, 600, "Imagens, vídeos, variações e testes."],
        ["Operação", "Reserva para reembolso", 100, 300, 1000, "Reserva evita quebrar por troca ou atraso."],
        ["Financeiro", "Taxas de pagamento", 20, 80, 200, "Processamento, antifraude e tarifas."],
        ["Legal", "Contabilidade", 0, 120, 500, "Validar necessidade com contador."],
        ["Legal", "CNPJ/MEI", 0, 200, 800, "Custos variam conforme formalização."],
        ["Ferramentas", "Ferramentas extras", 20, 100, 400, "Analytics, relatórios, automações e pesquisa."],
        ["Capital de giro", "Capital de giro", 300, 1200, 2000, "Cobertura para pedidos, prazo e contingência."],
    ]
    write_table(
        ws,
        ["Categoria", "Item", "Econômico", "Intermediário", "Profissional", "Observações"],
        rows,
        table_name="Tabela_Custos_Iniciais",
        table_style="TableStyleMedium7",
    )
    total_row = ws.max_row + 1
    ws.append(["", "Total estimado", f"=SUM(C2:C{total_row - 1})", f"=SUM(D2:D{total_row - 1})", f"=SUM(E2:E{total_row - 1})", "Soma das categorias"])
    for cell in ws[total_row]:
        cell.border = THIN_BORDER
        cell.fill = GOOD_FILL
        cell.font = Font(bold=True, color=PALETTE.green)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    apply_number_formats(ws, currency_cols=["C", "D", "E"], start_row=2)
    apply_sheet_finish(ws)


def build_pricing(wb: Workbook) -> None:
    ws = wb.create_sheet("07 Precificacao")
    ws.sheet_properties.tabColor = PALETTE.green
    write_title(
        ws,
        "Calculadora de preço, margem e CPA",
        "Campos azuis são editáveis. Campos amarelos são fórmulas protegidas contra alteração acidental. Ajuste as taxas conforme sua plataforma e valide imposto com contador.",
        end_col=7,
    )
    input_rows = [
        ["Custo do produto", 29.90, "Valor pago ao fornecedor por unidade"],
        ["Frete", 12.50, "Custo de envio médio por unidade"],
        ["Taxa da plataforma", 0.05, "Percentual sobre venda"],
        ["Taxa do pagamento", 0.032, "Percentual do gateway/antifraude"],
        ["Imposto estimado", 0.06, "Estimativa; validar com contador"],
        ["CPA/anúncio", 8.00, "Custo máximo por compra"],
        ["Reserva para reembolso", 3.50, "Valor reservado por pedido"],
        ["Margem desejada", 0.35, "Margem líquida alvo"],
    ]
    write_table(
        ws,
        ["Campo editável", "Valor", "Observação"],
        input_rows,
        start_row=4,
        table_name="Tabela_Precificacao_Inputs",
        table_style="TableStyleMedium2",
    )
    for row in range(5, 13):
        ws[f"B{row}"].fill = INPUT_FILL
        ws[f"B{row}"].font = Font(bold=True, color=PALETTE.navy)
        ws[f"B{row}"].protection = Protection(locked=False)
        ws[f"B{row}"].comment = Comment(
            f"Edite este valor para simular: {ws[f'A{row}'].value}.",
            "Codex",
        )
    for row in [5, 6, 10, 11]:
        ws[f"B{row}"].number_format = CURRENCY_FORMAT
    for row in [7, 8, 9, 12]:
        ws[f"B{row}"].number_format = PERCENT_FORMAT
    add_decimal_validation(
        ws,
        "B5:B6",
        minimum=0,
        maximum=10000,
        error="Use um valor em reais entre 0 e 10.000.",
    )
    add_decimal_validation(
        ws,
        "B7:B9",
        minimum=0,
        maximum=0.95,
        error="Use um percentual entre 0% e 95%.",
    )
    add_decimal_validation(
        ws,
        "B10:B11",
        minimum=0,
        maximum=10000,
        error="Use um valor em reais entre 0 e 10.000.",
    )
    add_decimal_validation(
        ws,
        "B12",
        minimum=0,
        maximum=0.95,
        error="Use uma margem desejada entre 0% e 95%.",
    )

    calculated_rows = [
        ["Custo fixo por pedido", "=B5+B6+B10+B11", "Produto + frete + CPA + reserva de reembolso"],
        ["Percentual consumido por taxas + margem", "=B7+B8+B9+B12", "Se chegar perto de 100%, a operação é inviável"],
        ["Preço mínimo", '=IF(B15>=1,"Rever taxas/margem",B14/(1-B15))', "Menor preço para atingir a margem desejada"],
        ["Preço sugerido", '=IF(ISNUMBER(B16),ROUNDUP(B16*1.05,0)-0.1,"Rever entradas")', "Preço com pequena folga comercial"],
        ["Receita líquida após taxas", '=IF(ISNUMBER(B17),B17*(1-B7-B8-B9),"")', "Venda menos taxas percentuais"],
        ["Custo total operacional", "=B14", "Custo fixo do pedido"],
        ["Lucro líquido por pedido", '=IF(ISNUMBER(B17),B18-B19,"")', "Resultado depois de produto, frete, CPA, reserva e taxas"],
        ["Margem percentual", '=IF(ISNUMBER(B17),B20/B17,"")', "Lucro líquido dividido pelo preço de venda"],
        ["ROI de anúncio", '=IF(B10>0,B20/B10,"")', "Lucro líquido dividido pelo CPA"],
    ]
    write_table(
        ws,
        ["Campo calculado", "Valor", "Leitura"],
        calculated_rows,
        start_row=14,
        table_name="Tabela_Precificacao_Resultados",
        table_style="TableStyleMedium9",
    )
    for row in range(15, 24):
        ws[f"B{row}"].fill = CALC_FILL
        ws[f"B{row}"].font = Font(bold=True, color=PALETTE.blue)
        ws[f"B{row}"].protection = Protection(locked=True)
    for row in [15, 17, 18, 19, 20]:
        ws[f"B{row}"].number_format = CURRENCY_FORMAT
    for row in [16, 22]:
        ws[f"B{row}"].number_format = PERCENT_FORMAT
    ws["B23"].number_format = ROI_FORMAT
    ws["A25"] = "Regra prática"
    ws["B25"] = (
        "Se o preço sugerido ficar fora do valor percebido pelo cliente ou a margem ficar abaixo de 20%, "
        "revise produto, fornecedor, frete, CPA ou oferta antes de anunciar."
    )
    ws.merge_cells("B25:G26")
    style_note(ws["A25"])
    style_note(ws["B25"])
    for row in range(25, 27):
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = THIN_BORDER
    add_numeric_alerts(ws, "B22")
    ws.protection.sheet = True
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = True
    apply_sheet_finish(ws, freeze="A4")
    set_column_widths(ws, widths={"A": 28, "B": 18, "C": 56, "D": 12, "E": 12, "F": 12, "G": 12})


def build_store_checklist(wb: Workbook) -> None:
    ws = wb.create_sheet("08 Loja")
    ws.sheet_properties.tabColor = PALETTE.blue
    rows = [
        ["Nome", "Marca e identificação", "Escolher nome claro, memorável e fácil de escrever", "Nome genérico ou difícil", "Pendente"],
        ["Logo", "Cria confiança", "Usar logo simples, legível e coerente com a categoria", "Logo poluído", "Pendente"],
        ["Paleta", "Ajuda a transmitir profissionalismo", "Usar 2 a 3 cores com contraste suficiente", "Cores fortes demais", "Pendente"],
        ["Domínio", "Facilita lembrança", "Escolher nome curto e sem caracteres estranhos", "Domínio longo", "Pendente"],
        ["Página inicial", "Define primeira impressão", "Mostrar categoria, benefício, produto e prova social honesta", "Página confusa", "Pendente"],
        ["Página de produto", "Aumenta conversão", "Foto, vídeo, descrição, preço, prazo e CTA visível", "Texto genérico", "Pendente"],
        ["Fotos", "Mostram o produto com realismo", "Usar imagens reais ou autorizadas e boa iluminação", "Foto copiada sem autorização", "Pendente"],
        ["Vídeos", "Aumentam confiança", "Mostrar uso real, antes/depois e escala do produto", "Vídeo sem contexto", "Pendente"],
        ["Descrição", "Explica benefício e uso", "Escrever problema, solução, conteúdo da embalagem e cuidados", "Promessa exagerada", "Pendente"],
        ["Prova social", "Reduz dúvida", "Usar avaliações reais e políticas transparentes", "Depoimento inventado", "Pendente"],
        ["Frete", "Evita frustração", "Informar prazo realista e rastreio", "Prazo curto sem base", "Pendente"],
        ["Política de troca", "Evita disputa", "Explicar prazo, condições e fluxo de atendimento", "Sem regra clara", "Pendente"],
        ["Política de privacidade", "Obrigatória para dados pessoais", "Descrever coleta, finalidade e contato", "Falta de transparência", "Pendente"],
        ["Termos de uso", "Protege operação", "Incluir regras da loja e limites de responsabilidade", "Sem termos", "Pendente"],
        ["Página de rastreio", "Melhora pós-venda", "Criar local claro para acompanhamento", "Cliente sem status", "Pendente"],
        ["WhatsApp ou e-mail", "Atendimento rápido", "Configurar canal, horário e resposta padrão", "Falta de resposta", "Pendente"],
        ["Checkout", "Fecha venda", "Reduzir etapas e campos desnecessários", "Checkout longo", "Pendente"],
        ["Teste de compra", "Valida processo inteiro", "Fazer pedido real e revisar e-mails, pagamento e rastreio", "Não testar", "Pendente"],
    ]
    write_table(
        ws,
        ["Item", "Por que é importante", "Como fazer bem", "Erro comum", "Status"],
        rows,
        table_name="Tabela_Loja",
    )
    add_status_validation(ws, f"E2:E{ws.max_row}")
    apply_sheet_finish(ws)


def build_paid_ads(wb: Workbook) -> None:
    ws = wb.create_sheet("09 Anuncios_Pagos")
    ws.sheet_properties.tabColor = "FF9896"
    rows = [
        ["Dias 1 a 7", "Meta Ads / TikTok Ads", "Pesquisa e criativos", "Teste de audiência inicial", "R$ 10 a R$ 20", "3 a 5 criativos por produto", "CTR, retenção e CPC", "CTR baixo ou CPC alto sem sinal de compra", "Não escalar nesta fase", "Validar promessa, ângulo e página antes de gastar mais."],
        ["Dias 8 a 14", "Meta Ads / TikTok Ads", "Testes pequenos", "Campanhas de 1 produto", "R$ 20 a R$ 40", "2 a 3 criativos vencedores", "CPA, checkout e conversão", "Sem venda após gasto relevante ao CPA máximo", "Escalar só se houver margem positiva", "Separar orçamento por produto e pausar rápido o que não responde."],
        ["Dias 15 a 21", "Meta Ads / TikTok Ads", "Otimização", "Ajuste de criativos, público e oferta", "R$ 30 a R$ 60", "Variações do melhor criativo", "ROAS e lucro líquido", "CPA acima do limite por 2 dias", "Aumentar 15% a 25% com dados", "Revisar frete, preço, prova social e checkout."],
        ["Dias 22 a 30", "Meta Ads / TikTok Ads", "Escala controlada", "Campanha com melhor performance", "R$ 40 a R$ 80", "Criativo principal + variações", "Lucro líquido e reclamações", "Queda de margem, atraso ou reclamação", "Escalar gradualmente mantendo reserva", "Não aumentar orçamento se fornecedor estiver instável."],
    ]
    write_table(
        ws,
        ["Dia/Semana", "Plataforma", "Objetivo", "Campanha", "Orçamento diário", "Criativos", "Métrica principal", "Regra para pausar", "Regra para escalar", "Observações"],
        rows,
        table_name="Tabela_Anuncios_Pagos",
        table_style="TableStyleMedium3",
    )
    apply_sheet_finish(ws)


def build_organic_traffic(wb: Workbook) -> None:
    ws = wb.create_sheet("10 Trafego_Organico")
    ws.sheet_properties.tabColor = PALETTE.purple
    rows = [
        ["TikTok", "5 a 7 vídeos/semana", "Antes e depois", "Mostre o problema, aplique o produto e finalize com o resultado", "Gerar alcance e salvar criativos vencedores", "Organizador de cabos", "Comente 'quero' ou acesse a loja"],
        ["Instagram Reels", "4 a 5 vídeos/semana", "Demonstração rápida", "Mostre uso real nos primeiros 3 segundos", "Construir confiança e remarketing", "Suporte adesivo para tomada", "Salve para testar em casa"],
        ["YouTube Shorts", "3 a 4 vídeos/semana", "Teste real", "Compare rotina antes/depois com legenda objetiva", "Gerar busca e prova de uso", "Escova removedora de pelos", "Veja mais detalhes na loja"],
        ["Pinterest", "2 a 3 pins/semana", "Organização visual", "Foto ou vídeo curto com resultado organizado", "Tráfego de cauda longa", "Organizador de geladeira", "Clique para ver a lista completa"],
        ["TikTok", "3 vídeos/semana", "3 coisas que resolvem...", "Liste três problemas e mostre uma solução prática", "Testar ângulos de comunicação", "Escova multifuncional", "Siga para mais soluções simples"],
        ["Instagram Reels", "2 vídeos/semana", "Cozinha prática", "Mostre preparo mais rápido ou espaço organizado", "Aumentar percepção de valor", "Escorredor dobrável", "Veja o produto em uso"],
        ["YouTube Shorts", "2 vídeos/semana", "Carro organizado", "Mostre interior bagunçado, aplicação e resultado", "Validar nicho automotivo leve", "Lixeira dobrável para carro", "Confira medidas e detalhes"],
        ["TikTok", "2 vídeos/semana", "Produto sendo usado por pet", "Mostre o pet, o problema e o uso sem promessa veterinária", "Validar apelo emocional", "Bebedouro portátil", "Comente se seu pet usaria"],
    ]
    write_table(
        ws,
        ["Plataforma", "Frequência", "Tipo de vídeo", "Roteiro curto", "Objetivo", "Produto ideal", "CTA"],
        rows,
        table_name="Tabela_Trafego_Organico",
        table_style="TableStyleMedium4",
    )
    apply_sheet_finish(ws)


def cronograma_rows() -> list[list[object]]:
    entries = [
        ("Definir nicho e posicionamento", "Pesquisar dores, concorrentes e demanda", "Lista de 10 ideias de produto", "Google Trends, TikTok, Instagram", "2h"),
        ("Selecionar produtos", "Filtrar produtos com utilidade clara e baixo risco", "5 produtos priorizados", "Planilha, marketplaces", "2h"),
        ("Escolher fornecedores", "Comparar Dropi, DSers, DropNacional e opções diretas", "Fornecedor preferido e alternativa", "Dropi, DSers, DropNacional", "2h"),
        ("Validar margem", "Calcular custo, frete, taxas, CPA máximo e reserva", "Tabela de margem por produto", "Aba Precificacao", "2h"),
        ("Comprar amostra", "Solicitar amostra do produto principal", "Pedido de amostra registrado", "Fornecedor", "1h"),
        ("Definir marca", "Nome, logo, paleta e tom de voz", "Branding básico", "Canva", "2h"),
        ("Montar estrutura da loja", "Criar conta, tema, menus e páginas principais", "Loja base pronta", "Nuvemshop/Shopify/Yampi", "3h"),
        ("Configurar pagamentos", "Ativar gateway, antifraude e dados de recebimento", "Checkout funcional", "Gateway de pagamento", "1h"),
        ("Criar políticas", "Troca, privacidade, termos e prazo de entrega", "Políticas publicadas", "Editor da loja", "2h"),
        ("Criar página de produto", "Texto, fotos, vídeo, preço, prazo e FAQ", "Página pronta", "Canva, CapCut, loja", "3h"),
        ("Criar criativos", "Produzir vídeos curtos e variações de gancho", "3 a 5 criativos", "CapCut", "3h"),
        ("Configurar pixel", "Instalar eventos e testar tracking", "Pixel funcionando", "Meta Ads, TikTok Ads", "1h"),
        ("Publicar conteúdo orgânico", "Postar primeiros vídeos e coletar sinais", "Conteúdo publicado", "TikTok, Reels, Shorts", "2h"),
        ("Criar campanha pequena", "Abrir teste com orçamento controlado", "Campanha ativa", "Meta Ads, TikTok Ads", "2h"),
        ("Analisar dados iniciais", "Ver CTR, CPC, retenção, checkout e conversão", "Resumo da primeira semana", "Analytics, Ads", "2h"),
        ("Ajustar criativos", "Trocar ganchos e formatos fracos", "Criativos novos", "CapCut", "2h"),
        ("Ajustar oferta", "Revisar CTA, preço, frete e prova social", "Oferta mais clara", "Loja", "2h"),
        ("Validar atendimento", "Testar WhatsApp, e-mail e respostas rápidas", "Fluxo de atendimento pronto", "WhatsApp Business", "1h"),
        ("Testar checkout", "Fazer pedido real e revisar todo o fluxo", "Checkout validado", "Loja", "1h"),
        ("Revisar logística", "Confirmar prazo, rastreio, embalagem e exceções", "Checklist de logística", "Fornecedor", "2h"),
        ("Criar reserva", "Separar reserva para reembolso e problemas", "Valor reservado", "Planilha financeira", "1h"),
        ("Escalar com cuidado", "Aumentar orçamento apenas no que tem margem", "Orçamento ajustado", "Anúncios", "2h"),
        ("Analisar margem real", "Comparar lucro real com projeção", "Relatório de margem", "Aba Financeiro", "2h"),
        ("Ajustar frete e prazo", "Revisar promessa de entrega e custo", "Frete ajustado", "Fornecedor, loja", "2h"),
        ("Criar mais conteúdo", "Publicar novos vídeos do melhor ângulo", "Novos conteúdos", "TikTok, Reels", "2h"),
        ("Revisar reputação", "Checar reclamações, mensagens e atrasos", "Plano de resposta", "Loja, atendimento", "2h"),
        ("Testar produto relacionado", "Adicionar item complementar com baixo risco", "Novo teste preparado", "Fornecedor, loja", "2h"),
        ("Avaliar pedidos", "Ver volume, prazo, satisfação e suporte", "Resumo operacional", "Analytics, planilha", "2h"),
        ("Repetir melhor criativo", "Escalar criativo vencedor com variações", "Criativo vencedor replicado", "CapCut, Ads", "2h"),
        ("Revisão final", "Decidir continuar, pausar ou trocar produto", "Decisão de continuidade", "Planilha e métricas", "2h"),
    ]
    return [[day, *entry, "Pendente"] for day, entry in enumerate(entries, start=1)]


def build_30_day_schedule(wb: Workbook) -> None:
    ws = wb.create_sheet("11 Cronograma_30_Dias")
    ws.sheet_properties.tabColor = "17BECF"
    write_table(
        ws,
        ["Dia", "Objetivo", "Tarefas", "Entregável do dia", "Ferramentas", "Tempo estimado", "Status"],
        cronograma_rows(),
        table_name="Tabela_Cronograma_30_Dias",
        table_style="TableStyleMedium2",
    )
    add_status_validation(ws, f"G2:G{ws.max_row}")
    apply_number_formats(ws, integer_cols=["A"], start_row=2)
    apply_sheet_finish(ws)


def build_risks(wb: Workbook) -> None:
    ws = wb.create_sheet("12 Riscos")
    ws.sheet_properties.tabColor = "8C564B"
    rows = [
        ["Atraso na entrega", "Fornecedor lento ou logística instável", "Cliente insatisfeito e pedido de reembolso", "Validar prazo realista e ter rastreio", "Comunicar rápido, oferecer solução e revisar fornecedor", "Alta", "Média"],
        ["Fornecedor ruim", "Baixa qualidade ou falta de estoque", "Reembolso, reclamação e perda de margem", "Comprar amostra e checar reputação", "Trocar fornecedor e pausar anúncio", "Alta", "Média"],
        ["Produto de baixa qualidade", "Material frágil ou acabamento ruim", "Devolução e reclamação pública", "Validar amostra e avaliações reais", "Substituir produto e atender cliente", "Alta", "Média"],
        ["Chargeback", "Pedido contestado ou atendimento ruim", "Perda financeira e risco no gateway", "Políticas claras, comprovantes e rastreio", "Responder com evidências e melhorar comunicação", "Alta", "Média"],
        ["Reembolso", "Cliente insatisfeito, atraso ou defeito", "Erosão de margem", "Manter reserva e prazo transparente", "Aplicar política de forma justa", "Média", "Alta"],
        ["Cliente insatisfeito", "Falta de informação ou expectativa errada", "Reclamação pública e queda de confiança", "Texto claro e atendimento em até 24h", "Responder com empatia e solução objetiva", "Alta", "Média"],
        ["Conta de anúncio bloqueada", "Promessa exagerada ou política violada", "Parada de tráfego", "Evitar claims sensíveis e revisar regras", "Reformular criativos e recorrer se aplicável", "Alta", "Baixa"],
        ["Produto saturado", "Concorrência alta e criativos iguais", "CPC alto e baixa conversão", "Testar ângulos e produtos relacionados", "Trocar oferta ou produto", "Média", "Média"],
        ["Margem baixa", "Frete, taxas e CPA acima do planejado", "Negócio sem lucro", "Calcular margem antes de anunciar", "Ajustar preço, oferta, fornecedor ou pausar", "Alta", "Alta"],
        ["Reclamação pública", "Atendimento lento ou promessa não cumprida", "Dano reputacional", "Responder rápido e documentar tudo", "Resolver com transparência e aprendizado", "Alta", "Média"],
        ["Problemas com importação", "Mercadoria internacional, taxa ou retenção", "Atraso, custo extra ou cancelamento", "Preferir fornecedor nacional no começo", "Usar alternativa nacional e informar cliente", "Alta", "Média"],
        ["Produto proibido", "Categoria restrita ou sem autorização", "Bloqueio, multa ou risco legal", "Validar categoria antes de vender", "Parar venda e buscar orientação", "Alta", "Baixa"],
        ["Falta de suporte", "Operação sem rotina", "Cliente sem resposta", "SLA e respostas rápidas", "Criar rotina diária e canal oficial", "Média", "Média"],
        ["Loja com aparência amadora", "Design, texto e checkout fracos", "Perda de confiança e baixa conversão", "Tema limpo, fotos reais e políticas claras", "Revisar página e teste de compra", "Média", "Alta"],
        ["Falta de capital de giro", "Crescimento sem reserva", "Pedidos atrasados e pausa forçada", "Separar caixa e reserva de reembolso", "Reduzir escala e reorganizar fluxo de caixa", "Alta", "Média"],
    ]
    write_table(
        ws,
        ["Problema", "Causa", "Consequência", "Como prevenir", "Como resolver", "Gravidade", "Probabilidade"],
        rows,
        table_name="Tabela_Riscos",
        table_style="TableStyleMedium3",
    )
    add_priority_validation(ws, f"F2:G{ws.max_row}")
    add_text_alerts(
        ws,
        f"F2:G{ws.max_row}",
        good_values=["Baixa"],
        warning_values=["Média"],
        bad_values=["Alta"],
    )
    apply_sheet_finish(ws)


def build_legal(wb: Workbook) -> None:
    ws = wb.create_sheet("13 Legal_Brasil")
    ws.sheet_properties.tabColor = PALETTE.blue
    rows = [
        ["CDC", "As regras do Código de Defesa do Consumidor se aplicam ao e-commerce.", "Ter políticas claras, atendimento e registros.", "Reclamações, sanções e perda de confiança.", "Validar com contador, advogado ou fonte oficial."],
        ["Direito de arrependimento", "Compras online têm regra de desistência em prazo legal.", "Informar procedimento de devolução e contato.", "Conflito com cliente e órgãos de defesa.", "Validar com fonte oficial."],
        ["Política de troca e devolução", "Cliente precisa entender prazos, condições e fluxo.", "Publicar política simples antes de vender.", "Reembolso improvisado e disputa.", "Validar com fonte oficial."],
        ["Transparência no prazo de entrega", "Prazo precisa ser claro e realista.", "Informar prazo conservador e rastreio.", "Insatisfação e reclamação.", "Validar com fonte oficial."],
        ["CNPJ/MEI", "A formalização depende da atividade, faturamento e estrutura.", "Consultar contador antes de escalar.", "Problemas fiscais e limite de operação.", "Validar com contador."],
        ["Nota fiscal", "A emissão depende do modelo fiscal aplicável.", "Entender obrigação antes de vender volume.", "Risco fiscal e dificuldade com fornecedores.", "Validar com contador."],
        ["Limite anual do MEI", "MEI tem limite de faturamento e atividades permitidas.", "Acompanhar receita mensal e projeção.", "Desenquadramento e imposto retroativo.", "Validar com contador ou fonte oficial."],
        ["Importação", "Importar pode ter imposto, documentação e retenção.", "Checar regra, prazo, rastreio e custo final.", "Atraso, custo extra e insatisfação.", "Validar com Receita Federal ou fonte oficial."],
        ["Remessa Conforme", "Compras internacionais podem exigir procedimentos específicos.", "Usar fornecedores com documentação e rastreio.", "Problemas na entrega e custo inesperado.", "Validar com fonte oficial."],
        ["Produtos proibidos", "Algumas categorias exigem autorização ou são restritas.", "Checar produto antes de cadastrar.", "Bloqueio, multa e risco legal.", "Validar com fonte oficial."],
        ["Promessas enganosas", "Benefícios exagerados geram risco legal e bloqueios.", "Usar comunicação honesta e demonstrável.", "Reclamações e bloqueio de anúncio.", "Validar políticas de anúncio e fonte oficial."],
        ["Uso de imagens do fornecedor", "Imagens podem ter direitos autorais ou restrições.", "Usar imagens próprias ou autorizadas.", "Notificação, remoção e disputa.", "Validar com fornecedor ou advogado."],
        ["LGPD", "Dados de clientes precisam de finalidade, cuidado e transparência.", "Ter aviso de privacidade e proteger acessos.", "Sanções e perda de confiança.", "Validar com fonte oficial ou jurídico."],
    ]
    write_table(
        ws,
        ["Tema", "Explicação simples", "O que fazer", "Risco se ignorar", "Fonte ou validação"],
        rows,
        table_name="Tabela_Legal_Brasil",
        table_style="TableStyleMedium2",
    )
    ws["A16"] = "Aviso"
    ws["B16"] = "Esta aba é orientação operacional básica, não aconselhamento jurídico definitivo."
    ws.merge_cells("B16:E16")
    style_note(ws["A16"])
    style_note(ws["B16"])
    apply_sheet_finish(ws)


def build_tools(wb: Workbook) -> None:
    ws = wb.create_sheet("14 Ferramentas")
    ws.sheet_properties.tabColor = PALETTE.gray
    rows = [
        ["Loja", "Nuvemshop", "Loja simples e rápida para Brasil", "R$ 29 a R$ 119/mês", "Sim", "Teste grátis ou plano inicial", "Sim", "Boa para começar com operação nacional."],
        ["Loja", "Shopify", "Loja completa e ecossistema amplo", "R$ 39 a R$ 399/mês", "Trial", "Nuvemshop", "Sim", "Boa para crescer, mas validar custo em reais."],
        ["Loja", "WooCommerce", "Loja com WordPress", "Variável", "Sim", "Nuvemshop", "Parcial", "Mais controle, exige manutenção."],
        ["Loja", "Yampi", "Checkout e estrutura de loja", "Variável", "Sim", "Nuvemshop", "Sim", "Muito usada no mercado brasileiro."],
        ["Domínio", "Registro.br ou registrador equivalente", "Comprar domínio próprio", "R$ 40 a R$ 80/ano", "Não", "Subdomínio temporário", "Sim", "Domínio próprio melhora confiança."],
        ["Design", "Canva", "Criar posts, banners e identidade", "Grátis / pago", "Sim", "Photopea", "Sim", "Excelente para MVP."],
        ["Vídeo", "CapCut", "Editar vídeos curtos", "Grátis / pago", "Sim", "VN", "Sim", "Útil para TikTok/Reels."],
        ["Anúncios", "Meta Ads", "Campanhas no Facebook e Instagram", "Orçamento variável", "Não", "Conteúdo orgânico", "Sim", "Começar com orçamento pequeno."],
        ["Anúncios", "TikTok Ads", "Campanhas na plataforma TikTok", "Orçamento variável", "Não", "Conteúdo orgânico", "Sim", "Bom para vídeos curtos."],
        ["Atendimento", "WhatsApp Business", "Atendimento, catálogo e respostas rápidas", "Grátis", "Sim", "E-mail", "Sim", "Canal essencial no Brasil."],
        ["Rastreio", "Melhor Envio", "Frete, rastreio e logística quando aplicável", "Grátis / variável", "Sim", "Rastreio manual", "Sim", "Útil para fornecedores nacionais."],
        ["E-mail marketing", "Brevo, Mailchimp ou similar", "Fluxos de pós-venda e recuperação", "Grátis / pago", "Sim", "E-mail da plataforma", "Parcial", "Só usar quando houver tráfego suficiente."],
        ["Analytics", "Google Analytics", "Medir tráfego e conversão", "Grátis", "Sim", "Relatórios da plataforma", "Sim", "Instalar desde o início."],
        ["Analytics", "Microsoft Clarity", "Mapas de calor e gravações", "Grátis", "Sim", "Hotjar grátis", "Sim", "Ajuda a achar problemas no checkout."],
        ["Pesquisa", "Google Trends", "Entender interesse e sazonalidade", "Grátis", "Sim", "Pesquisa manual", "Sim", "Bom para validar demanda."],
        ["Pesquisa de concorrentes", "Mercado Livre, Shopee, TikTok e Instagram", "Analisar preço, demanda e criativos", "Grátis", "Sim", "Pesquisa manual", "Sim", "Não copiar marca ou promessa."],
        ["Fornecedores", "Dropi", "Integração e catálogo", "Variável", "Validar", "Fornecedor direto", "Sim", "Validar plano no site oficial."],
        ["Fornecedores", "DSers", "Gestão de pedidos do AliExpress", "Grátis / pago", "Sim", "Dropi", "Sim", "Útil para testes internacionais."],
        ["Fornecedores", "CJdropshipping", "Sourcing e fulfillment", "Variável", "Validar", "Dropi", "Parcial", "Mais indicado após validação."],
        ["Automação", "Planilha + rotinas simples", "Controle inicial de pedidos e margem", "Grátis", "Sim", "Google Sheets", "Sim", "Automatizar só depois de validar."],
    ]
    write_table(
        ws,
        ["Categoria", "Ferramenta", "Função", "Preço aproximado", "Tem plano grátis?", "Alternativa gratuita", "Vale para iniciante?", "Observações"],
        rows,
        table_name="Tabela_Ferramentas",
        table_style="TableStyleMedium9",
    )
    add_flexible_yes_no_validation(ws, f"E2:E{ws.max_row}")
    add_flexible_yes_no_validation(ws, f"G2:G{ws.max_row}")
    add_text_alerts(
        ws,
        f"G2:G{ws.max_row}",
        good_values=["Sim"],
        warning_values=["Parcial"],
        bad_values=["Não"],
    )
    apply_sheet_finish(ws)


def build_financial_projection(wb: Workbook) -> None:
    ws = wb.create_sheet("15 Financeiro_3_Meses")
    ws.sheet_properties.tabColor = "BCBD22"
    write_title(
        ws,
        "Projeção financeira estimada para 3 meses",
        "Use como simulação. O resultado real depende de produto, criativo, fornecedor, tráfego, atendimento e execução.",
        end_col=15,
    )
    scenarios = {
        "Ruim": {
            "visitas": [1200, 1500, 1800],
            "taxa": [0.006, 0.008, 0.010],
            "ticket": [59, 62, 65],
            "ads_per_visit": 0.32,
            "tools": 80,
        },
        "Médio": {
            "visitas": [3000, 3800, 4600],
            "taxa": [0.014, 0.017, 0.020],
            "ticket": [75, 79, 84],
            "ads_per_visit": 0.28,
            "tools": 120,
        },
        "Bom": {
            "visitas": [5000, 6500, 8000],
            "taxa": [0.022, 0.026, 0.030],
            "ticket": [85, 90, 95],
            "ads_per_visit": 0.22,
            "tools": 180,
        },
    }
    rows: list[list[object]] = []
    row_num = 5
    for scenario_name, data in scenarios.items():
        for month_idx, month in enumerate(["Mês 1", "Mês 2", "Mês 3"]):
            rows.append(
                [
                    month,
                    scenario_name,
                    data["visitas"][month_idx],
                    data["taxa"][month_idx],
                    f"=ROUND(C{row_num}*D{row_num},0)",
                    data["ticket"][month_idx],
                    f"=E{row_num}*F{row_num}",
                    f"=G{row_num}*0.42",
                    f"=G{row_num}*0.10",
                    f"=G{row_num}*0.075",
                    f"=C{row_num}*{data['ads_per_visit']}",
                    data["tools"],
                    f"=G{row_num}*0.04",
                    f"=G{row_num}-H{row_num}-I{row_num}-J{row_num}-K{row_num}-L{row_num}-M{row_num}",
                    f'=IF(G{row_num}=0,0,N{row_num}/G{row_num})',
                ]
            )
            row_num += 1

    write_table(
        ws,
        [
            "Mês",
            "Cenário",
            "Visitas",
            "Taxa de conversão",
            "Pedidos",
            "Ticket médio",
            "Faturamento",
            "Custo dos produtos",
            "Frete",
            "Taxas",
            "Anúncios",
            "Ferramentas",
            "Reembolsos",
            "Lucro líquido",
            "Margem líquida",
        ],
        rows,
        start_row=4,
        table_name="Tabela_Financeiro_3_Meses",
        table_style="TableStyleMedium7",
    )
    apply_number_formats(
        ws,
        currency_cols=["F", "G", "H", "I", "J", "K", "L", "M", "N"],
        percent_cols=["D", "O"],
        integer_cols=["C", "E"],
        start_row=5,
    )
    data_end_row = 4 + len(rows)
    add_numeric_alerts(ws, f"O5:O{data_end_row}")
    ws.conditional_formatting.add(
        f"N5:N{data_end_row}",
        FormulaRule(
            formula=["=$N5<0"],
            fill=BAD_FILL,
            font=Font(color=PALETTE.red, bold=True),
        ),
    )
    ws.conditional_formatting.add(
        f"N5:N{data_end_row}",
        FormulaRule(
            formula=["=$N5>=0"],
            fill=GOOD_FILL,
            font=Font(color=PALETTE.green, bold=True),
        ),
    )

    summary_start = ws.max_row + 3
    summary_rows = []
    for scenario_idx, scenario in enumerate(["Ruim", "Médio", "Bom"], start=0):
        data_row = 7 + scenario_idx * 3
        summary_rows.append([scenario, f"=N{data_row}", f"=O{data_row}", f"=G{data_row}"])
    write_table(
        ws,
        ["Cenário", "Lucro líquido no mês 3", "Margem líquida no mês 3", "Faturamento no mês 3"],
        summary_rows,
        start_row=summary_start,
        table_name="Tabela_Financeiro_Resumo",
        table_style="TableStyleMedium4",
    )
    for row in range(summary_start + 1, summary_start + 4):
        ws[f"B{row}"].number_format = CURRENCY_FORMAT
        ws[f"C{row}"].number_format = PERCENT_FORMAT
        ws[f"D{row}"].number_format = CURRENCY_FORMAT
    add_numeric_alerts(ws, f"C{summary_start + 1}:C{summary_start + 3}")

    chart = LineChart()
    chart.style = 13
    chart.title = "Lucro líquido por cenário no mês 3"
    chart.y_axis.title = "R$"
    chart.height = 7
    chart.width = 12
    chart.add_data(Reference(ws, min_row=summary_start, max_row=summary_start + 3, min_col=2), titles_from_data=True)
    chart.set_categories(Reference(ws, min_row=summary_start + 1, max_row=summary_start + 3, min_col=1))
    ws.add_chart(chart, "F17")

    apply_sheet_finish(ws, freeze="A4")


def build_final_checklist(wb: Workbook) -> None:
    ws = wb.create_sheet("16 Checklist_Final")
    ws.sheet_properties.tabColor = PALETTE.green
    rows = [
        ["Nicho", "Nicho escolhido", "Sim", "Pendente", "Validar demanda, margem e risco."],
        ["Produto", "Produto validado", "Sim", "Pendente", "Checar utilidade, amostra e defeitos."],
        ["Fornecedor", "Fornecedor analisado", "Sim", "Pendente", "Confirmar prazo, estoque, rastreio e qualidade."],
        ["Validação", "Amostra comprada ou validada", "Sim", "Pendente", "Testar antes do anúncio pago."],
        ["Finanças", "Margem calculada", "Sim", "Pendente", "Usar a aba de precificação."],
        ["Loja", "Loja revisada", "Sim", "Pendente", "Checar páginas, imagens, CTA e confiança."],
        ["Checkout", "Checkout testado", "Sim", "Pendente", "Validar pagamento, e-mails e pedido."],
        ["Políticas", "Políticas criadas", "Sim", "Pendente", "Troca, privacidade, termos e prazo."],
        ["Prazo", "Prazo de entrega claro", "Sim", "Pendente", "Não prometer prazo curto sem base."],
        ["Atendimento", "Suporte configurado", "Sim", "Pendente", "WhatsApp e e-mail com rotina diária."],
        ["Criativos", "Criativos prontos", "Sim", "Pendente", "Vídeos curtos, claros e sem promessa exagerada."],
        ["Anúncios", "Anúncios configurados", "Sim", "Pendente", "Testes pequenos com CPA máximo definido."],
        ["Tracking", "Pixel instalado", "Sim", "Pendente", "Meta Pixel, TikTok Pixel e eventos quando aplicável."],
        ["Analytics", "Analytics instalado", "Sim", "Pendente", "Google Analytics ou alternativa."],
        ["Finanças", "Reserva para reembolso", "Sim", "Pendente", "Separar caixa antes de escalar."],
        ["Atendimento", "Plano de atendimento", "Sim", "Pendente", "Responder em até 24h úteis."],
        ["Conteúdo", "Plano de conteúdo", "Sim", "Pendente", "TikTok, Reels, Shorts e Pinterest se fizer sentido."],
        ["Escala", "Plano de escala", "Sim", "Pendente", "Aumentar orçamento apenas com margem positiva."],
    ]
    write_table(
        ws,
        ["Categoria", "Item", "Obrigatório?", "Status", "Observação"],
        rows,
        table_name="Tabela_Checklist_Final",
        table_style="TableStyleMedium4",
    )
    add_yes_no_validation(ws, f"C2:C{ws.max_row}")
    add_status_validation(ws, f"D2:D{ws.max_row}")
    apply_sheet_finish(ws)


def build_workbook() -> Workbook:
    wb = new_workbook()
    build_dashboard(wb)
    build_overview(wb)
    build_business_model(wb)
    build_niches(wb)
    build_products(wb)
    build_suppliers(wb)
    build_initial_costs(wb)
    build_pricing(wb)
    build_store_checklist(wb)
    build_paid_ads(wb)
    build_organic_traffic(wb)
    build_30_day_schedule(wb)
    build_risks(wb)
    build_legal(wb)
    build_tools(wb)
    build_financial_projection(wb)
    build_final_checklist(wb)
    if hasattr(wb, "calculation"):
        wb.calculation.fullCalcOnLoad = True
        wb.calculation.forceFullCalc = True
    wb.active = 0
    return wb


def create_preview_image(path: Path) -> bool:
    try:
        from PIL import Image, ImageDraw, ImageFont
    except Exception as exc:
        print(f"Preview image not generated: Pillow unavailable ({exc})")
        return False

    path.parent.mkdir(parents=True, exist_ok=True)

    width, height = 1400, 900
    img = Image.new("RGB", (width, height), color=(247, 250, 252))
    draw = ImageDraw.Draw(img)

    def font(size: int, bold: bool = False):
        candidates = [
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/liberation2/LiberationSans-Bold.ttf" if bold else "/usr/share/fonts/truetype/liberation2/LiberationSans-Regular.ttf",
        ]
        for candidate in candidates:
            if Path(candidate).exists():
                return ImageFont.truetype(candidate, size)
        return ImageFont.load_default()

    navy = (23, 54, 93)
    blue = (31, 78, 121)
    green = (46, 125, 50)
    purple = (103, 78, 167)
    amber = (192, 112, 33)
    gray = (80, 95, 112)
    border = (217, 226, 236)
    white = (255, 255, 255)

    draw.rounded_rectangle((36, 34, width - 36, height - 34), radius=18, fill=white, outline=border, width=2)
    draw.rectangle((36, 34, width - 36, 140), fill=navy)
    draw.text((70, 58), "Planejamento Dropshipping Brasil", fill=white, font=font(34, bold=True))
    draw.text(
        (72, 104),
        "Nicho recomendado, fornecedores, custos, precificação, riscos e projeção financeira",
        fill=(220, 234, 247),
        font=font(18),
    )

    cards = [
        ("Nicho recomendado", "Casa e organização", blue),
        ("Capital inicial", "R$ 400 a R$ 1.000", green),
        ("Fornecedor inicial", "Dropi ou nacional", purple),
        ("Canal orgânico", "TikTok + Reels", amber),
    ]
    x = 70
    for title, value, color in cards:
        draw.rounded_rectangle((x, 175, x + 285, 290), radius=14, fill=color)
        draw.text((x + 22, 198), title, fill=(235, 244, 252), font=font(15, bold=True))
        draw.text((x + 22, 235), value, fill=white, font=font(22, bold=True))
        x += 315

    draw.text((80, 345), "Ranking de nichos", fill=navy, font=font(24, bold=True))
    bars = [
        ("Casa e organização", 9.7, blue),
        ("Pet utilitário", 8.8, green),
        ("Cozinha prática", 8.5, amber),
        ("Automotivo leve", 8.3, purple),
        ("Fitness leve", 7.6, gray),
        ("Beleza não regulada", 7.4, gray),
    ]
    y = 392
    for label, score, color in bars:
        draw.text((85, y), label, fill=gray, font=font(16))
        draw.rounded_rectangle((310, y + 2, 760, y + 24), radius=8, fill=(232, 238, 244))
        draw.rounded_rectangle((310, y + 2, 310 + int(450 * score / 10), y + 24), radius=8, fill=color)
        draw.text((780, y - 1), f"{score:.1f}", fill=navy, font=font(17, bold=True))
        y += 46

    draw.text((870, 345), "Plano de execução", fill=navy, font=font(24, bold=True))
    steps = [
        "1. Validar 3 a 5 produtos simples",
        "2. Comprar amostra e testar prazo",
        "3. Criar loja limpa com políticas claras",
        "4. Publicar vídeos curtos antes de escalar",
        "5. Anunciar só com CPA e margem definidos",
        "6. Monitorar rastreio, suporte e reembolso",
    ]
    y = 392
    for step in steps:
        draw.rounded_rectangle((872, y - 8, 1285, y + 30), radius=8, fill=(243, 246, 249), outline=border)
        draw.text((895, y), step, fill=gray, font=font(16))
        y += 52

    draw.rounded_rectangle((70, 735, 1330, 820), radius=12, fill=(255, 242, 204), outline=(246, 178, 107))
    draw.text((95, 758), "Regra central", fill=amber, font=font(18, bold=True))
    draw.text(
        (95, 785),
        "Não é dinheiro fácil: escale apenas quando houver margem positiva, fornecedor estável, entrega previsível e atendimento sob controle.",
        fill=navy,
        font=font(18),
    )
    img.save(path)
    return True


def validate_workbook(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    wb = load_workbook(path, data_only=False)
    missing = [sheet for sheet in REQUIRED_SHEETS if sheet not in wb.sheetnames]
    if missing:
        raise RuntimeError(f"Missing required sheets: {', '.join(missing)}")

    if len(wb.sheetnames) != len(REQUIRED_SHEETS):
        raise RuntimeError(f"Unexpected sheet count: {len(wb.sheetnames)}")

    pricing = wb["07 Precificacao"]
    required_formulas = ["B15", "B16", "B17", "B20", "B22", "B23"]
    empty_formulas = [
        cell for cell in required_formulas
        if not isinstance(pricing[cell].value, str) or not pricing[cell].value.startswith("=")
    ]
    if empty_formulas:
        raise RuntimeError(f"Pricing formulas missing: {', '.join(empty_formulas)}")
    if not pricing.protection.sheet:
        raise RuntimeError("Pricing sheet protection is disabled.")
    if pricing["B5"].protection.locked:
        raise RuntimeError("Pricing input cells must remain editable.")

    costs = wb["06 Custos_Iniciais"]
    if not isinstance(costs[f"C{costs.max_row}"].value, str) or not costs[f"C{costs.max_row}"].value.startswith("="):
        raise RuntimeError("Initial cost total formula missing.")

    dashboard = wb["00 Escolha_Principal"]
    for cell_ref in ["F18", "F19", "F20", "I18", "I19", "I20"]:
        if not isinstance(dashboard[cell_ref].value, str) or not dashboard[cell_ref].value.startswith("="):
            raise RuntimeError(f"Dashboard linked formula missing in {cell_ref}.")

    finance = wb["15 Financeiro_3_Meses"]
    for cell_ref in ["E5", "G5", "N5", "O5"]:
        if not isinstance(finance[cell_ref].value, str) or not finance[cell_ref].value.startswith("="):
            raise RuntimeError(f"Financial projection formula missing in {cell_ref}.")

    for sheet_name in REQUIRED_SHEETS:
        ws = wb[sheet_name]
        if ws.max_row > 1 and not ws.auto_filter.ref and not ws.tables:
            raise RuntimeError(f"Sheet without filters/tables: {sheet_name}")

    products = wb["04 Produtos"]
    if not isinstance(products["J2"].value, str) or not products["J2"].value.startswith("="):
        raise RuntimeError("Product margin formula missing in Produtos!J2")
    for cell_ref in ["E2", "F2", "G2", "H2", "I2"]:
        if products[cell_ref].number_format != CURRENCY_FORMAT:
            raise RuntimeError(f"Currency format missing in Produtos!{cell_ref}")
    if products["J2"].number_format != PERCENT_FORMAT:
        raise RuntimeError("Percent format missing in Produtos!J2")
    for sheet_name in REQUIRED_SHEETS:
        ws = wb[sheet_name]
        if ws.freeze_panes is None:
            raise RuntimeError(f"Freeze panes missing in {sheet_name}")
        if ws.sheet_view.showGridLines:
            raise RuntimeError(f"Gridlines should be hidden in {sheet_name}")


def save_workbook(wb: Workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    validate_workbook(output_path)


def main() -> None:
    args = parse_args()
    wb = build_workbook()
    save_workbook(wb, args.output)

    preview_generated = False
    if not args.skip_preview:
        preview_generated = create_preview_image(args.preview)

    print(f"Workbook saved: {args.output}")
    print(f"Workbook validated: {args.output.exists()}")
    print(f"Sheets: {len(REQUIRED_SHEETS)}")
    if preview_generated:
        print(f"Preview saved: {args.preview}")
    elif not args.skip_preview:
        print("Preview saved: not generated")
        


if __name__ == "__main__":
    main()
